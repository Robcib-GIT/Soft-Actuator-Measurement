from Utilities.data_operations import load_data, save_data
import os
import numpy as np
from Utilities.blood_pressure import BloodPressure
from openpyxl import load_workbook

if __name__ == "__main__":

    carpeta = "data"
    all_data = {}

    for filename in os.listdir(carpeta):
        if filename.startswith("PRESION_") and filename.endswith(".csv"):
            filepath = os.path.join(carpeta, filename)
            data_dict, param_dict = load_data(filepath)
            # print(param_dict)

            fs = 1 / np.mean(np.diff(data_dict["Time"]))
            bp = BloodPressure(fs)

            # Obtener información
            subject = param_dict['subject']
            version = param_dict['version']

            results = {
                "real_sys": param_dict.get("SYS", None),
                "real_dia": param_dict.get("DIA", None),
                "real_ppm": param_dict.get("PPM", None)
            }

            # Calcular presiones con mi algoritmo
            try:
                sys, dia, ppm = bp.get_blood_pressure(data_dict["Pressure"])
                results["sam_sys"] = sys
                results["sam_dia"] = dia
                results["sam_ppm"] = ppm
            except Exception:
                results["sam_sys"] = None
                results["sam_dia"] = None
                results["sam_ppm"] = None

            # Almacenar informacion
            all_data.setdefault(subject, {})[version] = results

    # Ordenar la información por versions
    results_excel_dir = "data/RESULTADOS.xlsx"
    workbook = load_workbook(results_excel_dir)
    sheet1 = workbook["BloodPressure"]
    sheet2 = workbook["BP_Subject"]
    all_data_ordered = {}
    n = 0
    m = 0
    for subject, versions in all_data.items():
        sam_sys = []
        sam_dia = []
        sam_ppm = []
        real_sys = []
        real_dia = []
        real_ppm = []

        ordered_versions = dict(sorted(versions.items(), key=lambda item: item[0]))
        for version, results in ordered_versions.items():
            row = 3 + n
            sheet1.cell(row=row, column=1).value = n
            sheet1.cell(row=row, column=2).value = subject
            sheet1.cell(row=row, column=3).value = version

            if results['sam_sys'] is not None:
                sheet1.cell(row=row, column=4).value = results['sam_sys']
                sam_sys.append(results['sam_sys'])
            else:
                sheet1.cell(row=row, column=4).value = '=NA()'

            if results['sam_dia'] is not None:
                sheet1.cell(row=row, column=5).value = results['sam_dia']
                sam_dia.append(results['sam_dia'])
            else:
                sheet1.cell(row=row, column=5).value = '=NA()'

            if results['sam_ppm'] is not None:
                sheet1.cell(row=row, column=6).value = results['sam_ppm']
                sam_ppm.append(results['sam_ppm'])
            else:
                sheet1.cell(row=row, column=6).value = '=NA()'

            if results['real_sys'] is not None:
                sheet1.cell(row=row, column=7).value = results['real_sys']
                real_sys.append(results['real_sys'])
            else:
                sheet1.cell(row=row, column=7).value = '=NA()'

            if results['real_dia'] is not None:
                sheet1.cell(row=row, column=8).value = results['real_dia']
                real_dia.append(results['real_dia'])
            else:
                sheet1.cell(row=row, column=8).value = '=NA()'

            if results['real_ppm'] is not None:
                sheet1.cell(row=row, column=9).value = results['real_ppm']
                real_ppm.append(results['real_dia'])
            else:
                sheet1.cell(row=row, column=9).value = '=NA()'

            # sheet1.cell(row=row, column=4).value = results['sam_sys'] if results['sam_sys'] is not None else '=NA()'
            # sheet1.cell(row=row, column=5).value = results['sam_dia'] if results['sam_dia'] is not None else '=NA()'
            # sheet1.cell(row=row, column=6).value = results['sam_ppm'] if results['sam_ppm'] is not None else '=NA()'
            # sheet1.cell(row=row, column=7).value = results['real_sys'] if results['real_sys'] is not None else '=NA()'
            # sheet1.cell(row=row, column=8).value = results['real_dia'] if results['real_dia'] is not None else '=NA()'
            # sheet1.cell(row=row, column=9).value = results['real_ppm'] if results['real_ppm'] is not None else '=NA()'

            n += 1

        all_data_ordered[subject] = ordered_versions

        # Calcular medias y añadir
        mean_sam_sys = np.mean(sam_sys) if len(sam_sys) > 0 else None
        mean_sam_dia = np.mean(sam_dia) if len(sam_dia) > 0 else None
        mean_sam_ppm = np.mean(sam_ppm) if len(sam_ppm) > 0 else None
        mean_real_sys = np.mean(real_sys) if len(real_sys) > 0 else None
        mean_real_dia = np.mean(real_dia) if len(real_dia) > 0 else None
        mean_real_ppm = np.mean(real_ppm) if len(real_ppm) > 0 else None

        subject_row = 3 + m
        sheet2.cell(row=subject_row, column=1).value = subject
        sheet2.cell(row=subject_row, column=2).value = np.mean(sam_sys) if len(sam_sys) > 0 else '=NA()'
        sheet2.cell(row=subject_row, column=3).value = np.mean(sam_dia) if len(sam_dia) > 0 else '=NA()'
        sheet2.cell(row=subject_row, column=4).value = np.mean(sam_ppm) if len(sam_ppm) > 0 else '=NA()'
        sheet2.cell(row=subject_row, column=5).value = np.mean(real_sys) if len(real_sys) > 0 else '=NA()'
        sheet2.cell(row=subject_row, column=6).value = np.mean(real_dia) if len(real_dia) > 0 else '=NA()'
        sheet2.cell(row=subject_row, column=7).value = np.mean(real_ppm) if len(real_ppm) > 0 else '=NA()'

        m += 1

    # Guardar los cambios
    workbook.save(results_excel_dir)
